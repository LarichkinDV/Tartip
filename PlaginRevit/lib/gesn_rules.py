# -*- coding: utf-8 -*-
"""Работа с правилами подбора ГЭСН из Excel."""
import os
from collections import namedtuple

from openpyxl import load_workbook

from . import config

GesnRule = namedtuple(
    "GesnRule",
    [
        "family",
        "type_name",
        "thickness_mm",
        "height_min_mm",
        "height_max_mm",
        "reinforcement",
        "brick_size",
        "gesn_code",
        "unit_raw",
        "multiplier",
        "volume_param",
    ],
)


def _as_text(value):
    if value is None:
        return None
    try:
        return unicode(value)  # type: ignore[name-defined]
    except Exception:
        try:
            return str(value)
        except Exception:
            return None


def _as_float(value):
    try:
        if value is None:
            return None
        return float(value)
    except Exception:
        return None


def _normalize_bool_text(value):
    text = (_as_text(value) or u"").strip().lower()
    if not text:
        return u""
    return u"да" if text in {u"да", u"yes", u"1", u"true", u"истина"} else u"нет"


def load_rules_from_excel(path=None, sheet_name=None):
    """Загрузка правил из Excel.

    Возвращает список GesnRule. Пропускает пустые строки и строки без кода ГЭСН.
    """
    excel_path = path or config.EXCEL_PATH
    sheet = sheet_name or config.EXCEL_SHEET_NAME

    if not os.path.exists(excel_path):
        raise IOError(u"Файл правил не найден: {0}".format(excel_path))

    workbook = load_workbook(excel_path, data_only=True)
    if sheet not in workbook.sheetnames:
        raise ValueError(u"Лист '{0}' не найден в файле правил".format(sheet))

    ws = workbook[sheet]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {name: idx for idx, name in enumerate(header)}

    def get_cell(row, name):
        idx = header_map.get(name)
        if idx is None:
            return None
        return row[idx].value

    rules = []
    for row in ws.iter_rows(min_row=2):
        gesn_code = _as_text(get_cell(row, u"ГЭСН_код"))
        if not gesn_code:
            continue

        rule = GesnRule(
            family=_as_text(get_cell(row, u"Family")) or u"",
            type_name=_as_text(get_cell(row, u"TypeName")) or u"",
            thickness_mm=_as_float(get_cell(row, u"Thickness_mm")) or 0.0,
            height_min_mm=_as_float(get_cell(row, u"Height_min_mm")) or 0.0,
            height_max_mm=_as_float(get_cell(row, u"Height_max_mm")) or 0.0,
            reinforcement=_normalize_bool_text(get_cell(row, u"Армирование")),
            brick_size=_as_text(get_cell(row, u"Размеры кирпича")) or u"",
            gesn_code=gesn_code,
            unit_raw=_as_text(get_cell(row, u"ЕдИзм")) or u"",
            multiplier=_as_float(get_cell(row, u"Кратность")) or 1.0,
            volume_param=_as_text(get_cell(row, u"Параметр_объёма")) or u"",
        )
        rules.append(rule)

    return rules
